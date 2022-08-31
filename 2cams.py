import datetime
import cv2
import numpy as np
import os
import time
#from imutils.video import FPS
import pandas as pd
#import schedule
from openpyxl import load_workbook

def read_config_cams():
    # читаем конфиг
    global input1
    global input2
    check_config = os.path.exists('config.ini')
    if (check_config == True):
        f = open('config.ini', 'r')
        for line in f:
            l = [line.strip() for line in f]
        f.close()
        input1 = f'{l[0]}'
        input2 = f'{l[1]}'
        #print(input1)
        #print(input2)
        return input1, input2
    else:
        print('файла не существует, создаем')
        lines = ["# Настройка камер", "0", "0"]
        with open(r"config.ini", "w") as file:
            for line in lines:
                file.write(line + '\n')


# Создаем БД в эксель по умолчанию


def db_test():
    default_dir_db = 'db'
    person_in_fio = 'Не определено'
    person_in_time = 'Время не определено'
    person_out_fio = 'Не определено'
    person_out_time = 'Время не определено'
    sheet_name1 = datetime.datetime.now()
    sheet_name1 = sheet_name1.strftime("%Y-%m-%d")
    os.makedirs(f'{default_dir_db}/{sheet_name1}', exist_ok=True)
    db_default = 'db.xlsx'
    # Создаем пустой датафрейм за текущую дату если отсутствует
    check_file = os.path.exists(f'{default_dir_db}/{sheet_name1}/db.xlsx')

    if (check_file == False):
        db_1 = pd.DataFrame(columns=['id_dt', 'event', 'id_fio', 'fio', 'fio_path', 'status'])
        # Создаем пустой excel на текущую дату если она не существует
        db_1.to_excel(f'{default_dir_db}/{sheet_name1}/db.xlsx', sheet_name=sheet_name1, index=False)
        temp_db_in = db_1.copy()
        #temp_db_in.drop(temp_db_in.index, inplace=True)
        temp_db_out = db_1.copy()
        #temp_db_out.drop(temp_db_out.index, inplace=True)
    else:
        db_1 = pd.read_excel(f'{default_dir_db}/{sheet_name1}/db.xlsx', sheet_name=sheet_name1)
        filter_in = db_1['status'] == 'in'
        temp_db_in = db_1.loc[filter_in].copy()
        filter_out = db_1['status'] == 'out'
        temp_db_out = db_1.loc[filter_out].copy()
        if len(temp_db_in.index) > 0:
            person_in_fio = temp_db_in['fio'].where(temp_db_in['id_dt'] == temp_db_in['id_dt'].max()).dropna().values[
                0]
            person_in_time = temp_db_in['id_dt'].where(
                temp_db_in['id_dt'] == temp_db_in['id_dt'].max()).dropna().tolist()
            person_in_time = person_in_time[0]
            person_in_time = datetime.datetime.strftime(person_in_time, '%Y-%m-%d %H:%M:%S')
            temp_db_in.drop(temp_db_in.index, inplace=True)
        if len(temp_db_out.index) > 0:
            person_out_fio = \
                temp_db_out['fio'].where(temp_db_out['id_dt'] == temp_db_out['id_dt'].max()).dropna().values[0]
            person_out_time = temp_db_out['id_dt'].where(
                temp_db_out['id_dt'] == temp_db_out['id_dt'].max()).dropna().tolist()
            person_out_time = person_out_time[0]
            person_out_time = datetime.datetime.strftime(person_out_time, '%Y-%m-%d %H:%M:%S')
            temp_db_out.drop(temp_db_out.index, inplace=True)

    return person_in_fio, \
           person_in_time, \
           person_out_fio, \
           person_out_time, \
           temp_db_in, temp_db_out,\
           default_dir_db, \
           sheet_name1

def pred_init():
    datasets = 'datasets'
    global model
    global face_cascade
    (images, lables, names, id) = ([], [], {}, 0)
    for (subdirs, dirs, files) in os.walk(datasets):
        for subdir in dirs:
            names[id] = subdir
            subjectpath = os.path.join(datasets, subdir)
            for filename in os.listdir(subjectpath):
                path = subjectpath + '/' + filename
                lable = id
                images.append(cv2.imread(path, 0))
                lables.append(int(lable))
            #print(images)
            id += 1

    (images, lables) = [np.array(lis) for lis in [images, lables]]
    model = cv2.face.LBPHFaceRecognizer_create()
    model.train(images, lables)
    haar_file = 'haarcascade_frontalface_default.xml'
    face_cascade = cv2.CascadeClassifier(haar_file)
    #print(id)
    return images, \
           lables, \
           names, \
           id


def db_apply(temp_db):
    default_dir_db = db_test()[6]
    sheet_name1 = db_test()[7]

    #filter_in = temp_db_in['status'] == 'in'
    #temp_db_in = temp_db_in.loc[temp_db_in.groupby("id_fio")['id_dt'].idxmax()]

    #filter_out = temp_db_out['status'] == 'out'
    #temp_db_out = temp_db_out.loc[temp_db_out.groupby("id_fio")['id_dt'].idxmax()]

    #col_names = ['id_dt_in', 'event_in', 'id_fio_in', 'fio_in', 'fio_path', 'status']
    # Сбор данных во временную бд
    #temp_db_in.to_excel(f'{default_dir_db}/{sheet_name1}/temp_db_in.xlsx', sheet_name=sheet_name1, index=False)
    #temp_db_out.to_excel(f'{default_dir_db}/{sheet_name1}/temp_db_out.xlsx', sheet_name=sheet_name1,
    #                     index=False)

    book = load_workbook(f'{default_dir_db}/{sheet_name1}/db.xlsx')
    writer = pd.ExcelWriter(f'{default_dir_db}/{sheet_name1}/db.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    for sheetname in writer.sheets:
        temp_db.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row,
                            index=False,
                            header=False)
    writer.save()
    temp_db.drop(temp_db.index, inplace=True)
    #os.remove(f'{default_dir_db}/{sheet_name1}/temp_db_in.xlsx')
    #os.remove(f'{default_dir_db}/{sheet_name1}/temp_db_out.xlsx')

def statusin(id_in):
    default_dir_db = db_test()[6]
    sheet_name1 = db_test()[7]
    status_in = False
    #print(id_in)
    check_file = os.path.exists(f'{default_dir_db}/{sheet_name1}/db.xlsx')
    if (check_file == False):
        #бд за сегодня пустая, status_in=0
        status_in = False
    else:
        db_1 = pd.read_excel(f'{default_dir_db}/{sheet_name1}/db.xlsx', sheet_name=sheet_name1)
        filter_in = db_1['id_fio'] == id_in
        temp_db_in = db_1.loc[filter_in].copy()
        temp_db_in2 = temp_db_in['status'].where(temp_db_in['id_dt'] == temp_db_in['id_dt'].max()).dropna().values[0]
        #print(temp_db_in2)
        if (temp_db_in2 == 'in'):
            status_in = True
        else:
            status_in = False

    return status_in

def statusout(id_out):
    default_dir_db = db_test()[6]
    sheet_name1 = db_test()[7]
    status_out = False
    #print(id_out)
    check_file = os.path.exists(f'{default_dir_db}/{sheet_name1}/db.xlsx')
    if (check_file == False):
        #бд за сегодня пустая, status_out=0
        status_out = False
    else:
        db_1 = pd.read_excel(f'{default_dir_db}/{sheet_name1}/db.xlsx', sheet_name=sheet_name1)
        filter_out = db_1['id_fio'] == id_out
        temp_db_out = db_1.loc[filter_out].copy()
        temp_db_out2 = temp_db_out['status'].where(temp_db_out['id_dt'] == temp_db_out['id_dt'].max()).dropna().values[0]
        print(temp_db_out2)
        if (temp_db_out2 == 'out'):
            status_out = True
        else:
            status_out = False

    return status_out

def try_cams():

    #rint(db_test()[0], db_test()[1], db_test()[2], db_test()[3])
    #print(read_config_cams()[0]+'5')
    pif = db_test()[0]
    pit = db_test()[1]
    pof = db_test()[2]
    pot = db_test()[3]
    temp_db_in = db_test()[4]
    temp_db_out = db_test()[5]
    (width, height) = (130, 100)

    try:
        #webcamin = cv2.VideoCapture(path)
        webcamin = cv2.VideoCapture(read_config_cams()[0])
        webcamout = cv2.VideoCapture(read_config_cams()[1])

        #webcamin = cv2.VideoCapture(0, cv2.CAP_DSHOW)
  
        #print(read_config_cams()[0])
        #print(read_config_cams()[1])
    except:
        print('Не удалось запустить поток')
        quit()
    while True:

        (retin, imin) = webcamin.read()
        if (retin):
            names = pred_init()[2]
            #ids=names.keys()
            #print(ids)
            grayin = cv2.cvtColor(imin, cv2.COLOR_BGR2GRAY)
            facesin = face_cascade.detectMultiScale(grayin, 1.2, 3)
            for (x, y, w, h) in facesin:
                cv2.rectangle(imin, (x, y), (x + w, y + h), (255, 0, 0), 2)
                facein = grayin[y:y + h, x:x + w]
                face_resize_in = cv2.resize(facein, (width, height))
                predictionin = model.predict(face_resize_in)
                cv2.rectangle(imin, (x, y), (x + w, y + h), (0, 255, 0), 3)
                if ((predictionin[1] < 90) and (statusin(predictionin[0]) == False)):
                    cv2.putText(imin, 'The person of % s - %.0f' % (names[predictionin[0]], predictionin[1]),
                                (x - 10, y - 10),
                                cv2.FONT_HERSHEY_PLAIN, 1, (0, 0, 255), 2)
                    now1 = datetime.datetime.now()
                    # now1=datetime.datetime.strftime(now1, '%Y-%m-%d %H:%M:%S')
                    id_dt_in = now1
                    event_in = 'Enter'
                    id_fio_in = predictionin[0]
                    # print(predictionin[1])
                    fio_in = f'{names[predictionin[0]]}'
                    fio_path = f'datasets/{names[predictionin[0]]}'
                    status = 'in'
                    # print(id_dt_in, event_in, id_fio_in, fio_in, fio_path, status)
                    new_row = {'id_dt': now1,
                               'event': event_in,
                               'id_fio': id_fio_in,
                               'fio': fio_in,
                               'fio_path': fio_path,
                               'status': status
                               }
                    temp_db_in.loc[len(temp_db_in.index)] = new_row
                    #print(temp_db_in)
                    pif = temp_db_in['fio'].where(temp_db_in['id_dt'] == temp_db_in['id_dt'].max()).dropna().values[0]
                    print(pif)
                    pit = temp_db_in['id_dt']. \
                        where(temp_db_in['id_dt'] == temp_db_in['id_dt'].max()).dropna().tolist()
                    pit = pit[0]
                    pit = datetime.datetime.strftime(pit, '%Y-%m-%d %H:%M:%S')
                    db_apply(temp_db_in)
                    temp_db_in.drop(temp_db_in.index, inplace=True)

            info = [
                ("ФИО последнего вошедшего", pif),
                ("Время входа", pit),
                ("Текущее время", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            ]
            for (i, (k, v)) in enumerate(info):
                text = "{}: {}".format(k, v)
                cv2.putText(imin, text, (8, ((i * 20) + 20)),
                            cv2.FONT_HERSHEY_COMPLEX, 0.6, (0, 255, 0), 2)
            cv2.imshow('Livein', imin)

        (retout, imout) = webcamout.read()
        if (retout):
            grayout = cv2.cvtColor(imout, cv2.COLOR_BGR2GRAY)
            facesout = face_cascade.detectMultiScale(grayout, 1.3, 4)
            for (x, y, w, h) in facesout:
                cv2.rectangle(imout, (x, y), (x + w, y + h), (255, 0, 0), 2)
                faceout = grayout[y:y + h, x:x + w]
                face_resize_out = cv2.resize(faceout, (width, height))
                predictionout = model.predict(face_resize_out)
                cv2.rectangle(imout, (x, y), (x + w, y + h), (0, 0, 255), 3)
                if ((predictionout[1] < 90) and (statusout(predictionout[0]) == False)):
                    cv2.putText(imout, ('The person of % s - %.0f' % (names[predictionout[0]], predictionout[1])),
                                (x - 10, y - 10),
                                cv2.FONT_HERSHEY_PLAIN, 1, (0, 0, 255), 2)
                    now1 = datetime.datetime.now()
                    # now1=datetime.datetime.strftime(now1, '%Y-%m-%d %H:%M:%S')
                    id_dt_out = now1
                    event_out = 'Exit'
                    id_fio_out = predictionout[0]
                    # print(predictionout[0])
                    fio_out = f'{names[predictionout[0]]}'
                    fio_path = f'datasets/{names[predictionout[0]]}'
                    status = 'out'
                    # print(id_dt_out, event_out, id_fio_out, fio_out, fio_path, status)
                    new_row = {'id_dt': now1,
                               'event': event_out,
                               'id_fio': id_fio_out,
                               'fio': fio_out,
                               'fio_path': fio_path,
                               'status': status
                               }
                    temp_db_out.loc[len(temp_db_out.index)] = new_row

                    pof = \
                        temp_db_out['fio'].where(temp_db_out['id_dt'] == temp_db_out['id_dt'].max()).dropna().values[0]
                    pot = temp_db_out['id_dt'].where(
                        temp_db_out['id_dt'] == temp_db_out['id_dt'].max()).dropna().tolist()
                    pot = pot[0]
                    pot = datetime.datetime.strftime(pot, '%Y-%m-%d %H:%M:%S')
                    db_apply(temp_db_out)
                    temp_db_out.drop(temp_db_in.index, inplace=True)
            info = [
                ("ФИО последнего вышедшего", pof),
                ("Время выхода", pot),
                ("Текущее время", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            ]
            for (i, (k, v)) in enumerate(info):
                text = "{}: {}".format(k, v)
                cv2.putText(imout, text, (8, ((i * 20) + 20)),
                            cv2.FONT_HERSHEY_COMPLEX, 0.6, (0, 0, 255), 2)
            cv2.imshow('LiveOut', imout)
        key = cv2.waitKey(1)
        if ((key == 27) or (retout == False) or (retin == False)):
            webcamin.release()
            webcamout.release()
            cv2.destroyAllWindows()

            break

def main():
    db_test()
    pred_init()
    try_cams()


if __name__ == '__main__':

    main()

